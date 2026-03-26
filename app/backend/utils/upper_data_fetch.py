import requests
from urllib.parse import quote
from app.backend.config import UPPER_AIR_DATA_DIR 
from app.backend.utils.ogimet import OgimetAPI
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from datetime import datetime
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import time
import pandas as pd
from PyPDF2 import PdfReader




def fetch_upper_air_data(datetime_str: str, station_id: str, src: str = 'UNKNOWN', data_type: str = 'TEXT:CSV') -> str:
    """
    Fetch upper air sounding data from University of Wyoming's weather site.

    Args:
        datetime_str (str): DateTime in format "YYYY-MM-DD HH:MM:SS"
        station_id (str): 5-digit WMO station ID (e.g. "43003")
        src (str): Source (default is 'UNKNOWN')
        data_type (str): Format type (default is 'TEXT:CSV')

    Returns:
        str: Raw upper air data as plain text

    Raises:
        Exception: If data not available or fetch fails
    """
    base_url = "https://weather.uwyo.edu/wsgi/sounding"
    datetime_encoded = quote(datetime_str)
    full_url = f"{base_url}?datetime={datetime_encoded}&id={station_id}&src={src}&type={data_type}"


    session = requests.Session()
    retries = Retry(
        total=5,  # Increased from 3 to 5 retries
        backoff_factor=2,  # Increased exponential backoff from 1 to 2
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset(["GET"])
    )
    session.mount("https://", HTTPAdapter(max_retries=retries))
    session.mount("http://", HTTPAdapter(max_retries=retries))

    # Increased timeout from 10s to 30s (connect_timeout, read_timeout)
    timeout = (10, 30)  # (connect_timeout, read_timeout)
    max_retries_manual = 3
    last_error = None
    
    for attempt in range(max_retries_manual):
        try:
            print(f"[DEBUG] Fetching upper-air from: {full_url} (attempt {attempt + 1}/{max_retries_manual})")
            response = session.get(full_url, timeout=timeout)
            print(f"[DEBUG] Upper-air fetch succeeded with status {response.status_code}")
            break
        except requests.exceptions.Timeout as e:
            last_error = e
            print(f"[WARN] Upper-air fetch timeout (attempt {attempt + 1}/{max_retries_manual}): {e}")
            if attempt < max_retries_manual - 1:
                wait_time = 2 ** attempt  # 1s, 2s, 4s
                print(f"[INFO] Retrying after {wait_time}s...")
                time.sleep(wait_time)
                continue
            raise Exception(f"Failed to fetch upper-air data after {max_retries_manual} attempts: {e}")
        except requests.exceptions.RequestException as e:
            # Connection refused, DNS problems, etc.
            last_error = e
            print(f"[ERROR] Failed to fetch upper-air data: {e}")
            raise Exception(f"Failed to fetch upper-air data: {e}")

    if response.status_code != 200:
        raise Exception(f"Failed to fetch data. HTTP Status Code: {response.status_code}")

    if '<html>' in response.text.lower():
        raise Exception("HTML page received: likely no data available for this datetime/station.")

    # Save to file
    download_dir = os.path.join(UPPER_AIR_DATA_DIR, 'downloads')
    os.makedirs(download_dir, exist_ok=True)
    dt = datetime_str.replace(":", "").replace("-", "").replace(" ", "_")
    filename = secure_filename(f"upper_air_{station_id}_{dt}.csv")
    file_path = os.path.join(download_dir, filename)
    print(f"[DEBUG] Saving upper-air to: {file_path}")
    with open(file_path, "w", encoding="utf-8") as f:
        f.write(response.text)
    print(f"[INFO] Upper-air data saved to {file_path}")
    return file_path


def interpolate_temperature_only(actual_df, forecast_df):
    results = []

    for _, forecast_row in forecast_df.iterrows():
        forecast_alt = forecast_row["Altitude (m)"]

        # Get two actual levels above and below
        below = actual_df[actual_df["geopotential height_m"] <= forecast_alt]
        above = actual_df[actual_df["geopotential height_m"] >= forecast_alt]

        if below.empty or above.empty:
            continue  # Skip if interpolation not possible

        lower = below.iloc[-1]
        upper = above.iloc[0]

        h1, h2 = lower["geopotential height_m"], upper["geopotential height_m"]
        print(f"[DEBUG] Lower level: {h1} m, Upper level: {h2} m for forecast altitude {forecast_alt} m")
        t1, t2 = lower["temperature_C"], upper["temperature_C"]

        # Interpolate temperature
        if h2 - h1 != 0:
            interp_temp = ((h2 - forecast_alt) * t1 + (forecast_alt - h1) * t2) / (h2 - h1)
        else:
            interp_temp = t1
        print(f"[DEBUG] Interpolated temperature at {forecast_alt} m: {interp_temp:.2f} C")

        # For other parameters, take the closer one (nearest actual level)
        if abs(h1 - forecast_alt) <= abs(h2 - forecast_alt):
            nearest_row = lower
        else:
            nearest_row = upper

        results.append({
            **forecast_row.to_dict(),
            "interp_temperature_C": interp_temp,
            "actual_wind_speed_m/s": nearest_row["wind speed_m/s"],
            "actual_wind_direction": nearest_row.get("wind direction_degree")
        })

    return pd.DataFrame(results)

def parse_forecast_pdf(pdf_path):
    reader = PdfReader(pdf_path)
    text = "\n".join(page.extract_text() for page in reader.pages)

    match = re.search(r"UPPER WINDS(.*?)WEATHER", text, re.DOTALL)
    if not match:
        raise ValueError("Upper Winds section not found in PDF.")
    upper_winds_text = match.group(1)

    icaoM = re.search(r"LOCAL FORECAST FOR(.*?)AND", text)
    if not icaoM:
        raise ValueError("ICAO code not found in PDF.")
    icao = icaoM.group(1).strip()
    print(f"ICAO code extracted: {icao}")

    startDateTimeM = re.search(r"FROM(.*?)UTC", text,re.DOTALL)
    if not startDateTimeM:
        raise ValueError("Start date and time not found in PDF.")
    
    startDateTimeRaw = startDateTimeM.group(1).strip()
    try:
    # Example: if the PDF gives "01 Jan 2023 00:00"
        dt = datetime.strptime(startDateTimeRaw, "%Y/%m/%d %H:%M")
        startDateTime = dt.strftime("%Y%m%d%H%M")
        start_hour = dt.strftime("%H")
    except ValueError:
        raise ValueError(f"Could not parse date/time: '{startDateTimeRaw}'")

    endDateTimeM = re.search(r"TO(.*?)UTC", text,re.DOTALL)
    if not endDateTimeM:
        raise ValueError("Start date and time not found in PDF.")
    
    endDateTimeRaw = endDateTimeM.group(1).strip()
    try:
    # Example: if the PDF gives "01 Jan 2023 00:00"
        dt = datetime.strptime(endDateTimeRaw, "%Y/%m/%d %H:%M")
        endDateTime = dt.strftime("%Y%m%d%H%M")
        end_hour = dt.strftime("%H")
    except ValueError:
    # Try another format if needed, or raise error
        raise ValueError(f"Could not parse date/time: '{endDateTimeRaw}'")
    
    validity_code = f"{start_hour}-{end_hour}"


    # Extract WEATHER section (from 'WEATHER' to end or next section)
    weather_match = re.search(r"WEATHER(.*?)(?==)", text, re.DOTALL)
    weather_text = weather_match.group(1).strip() if weather_match else ""

    # Extract wind data
    pattern = re.findall(r"(\d+)[Mm]\s+(\d{3})/(\d{2})\s+([+-]?\d{2})", upper_winds_text)
    data = [(int(alt), dir, speed, temp) for alt, dir, speed, temp in pattern]
    data.sort(reverse=True)

    df = pd.DataFrame(data, columns=["Altitude (m)", "Wind Direction", "Wind Speed (kt)", "Temperature (°C)"])
    return df, weather_text,startDateTime, endDateTime,icao,validity_code

def validate_forecast_weather_with_metar(forecast_pdf_path):
    """
    Validates forecast weather condition against actual METARs for the same time range.
    """
    try:
        forecast_df, forecast_weather, start_time, end_time, icao, _ = parse_forecast_pdf(forecast_pdf_path)
        print(f"[INFO] Forecast weather: {forecast_weather}")
        print(f"[INFO] ICAO: {icao}, Time: {start_time} to {end_time}")

        # Fetch METAR using Ogimet
        api = OgimetAPI()
        metar_file_path = api.save_metar_to_file(begin=start_time, end=end_time, icao=icao)

        if not os.path.exists(metar_file_path):
            raise FileNotFoundError(f"METAR file not found: {metar_file_path}")

        with open(metar_file_path, 'r', encoding='utf-8') as f:
            metar_lines = f.readlines()

        forecast_keywords = re.findall(r'\b[A-Z]{2,}\b', forecast_weather)
        forecast_keywords = [w.strip() for w in forecast_keywords if w.isalpha()]
        print(f"[DEBUG] Forecast weather keywords: {forecast_keywords}")

        found_keywords = []
        for line in metar_lines:
            for keyword in forecast_keywords:
                if keyword in line:
                    found_keywords.append(keyword)

        match_status = "CORRECT" if found_keywords else "INCORRECT"
        match_percentage = 100 if found_keywords else 0

        return {
            "status": match_status,
            "metar_lines": metar_lines,
            "match_percentage": match_percentage,
            "matched_keywords": list(set(found_keywords)),
            "forecast_text": forecast_weather,
        }

    except Exception as e:
        print(f"[ERROR] Weather verification failed: {e}")
        return {
            "status": "ERROR",
            "match_percentage": 0,
            "error": str(e),
            "metar_lines": [],
            "matched_keywords": []
        }


def generate_upper_air_verification_xlsx(data_rows, metadata, file_path, weather_info=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Upper Air Verification"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def is_merged(ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if (row >= merged_range.min_row and row <= merged_range.max_row and
                col >= merged_range.min_col and col <= merged_range.max_col):
                return (merged_range.min_row, merged_range.min_col) != (row, col)
        return False

    def write_cell(ws, row, col, value, bold=False, center=True):
        if is_merged(ws, row, col):
            return
        cell = ws.cell(row=row, column=col)
        cell.value = value
        if bold:
            cell.font = Font(bold=True)
        if center:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Push table down by 2 rows for the heading
    ws.insert_rows(1, amount=2)

    # --- Bilingual Heading ---
    if data_rows:
        first_date = data_rows[0].get("date", "")
        try:
            dt_obj = datetime.strptime(first_date, "%d/%m/%Y")
            month_name = dt_obj.strftime("%B")
            year_val = dt_obj.year
        except:
            month_name = "Unknown"
            year_val = ""

        heading_en = f"Verification of Local / Area Forecasts for the month of {month_name} - {year_val} in r/o AMO Mumbai"
        heading_hi = f"{month_name},{year_val} के लिए लोकल /एरिया पूर्वानुमान का सत्यापन रिपोर्ट"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
        ws.cell(row=1, column=1, value=heading_en).font = Font(bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
        ws.cell(row=2, column=1, value=heading_hi).font = Font(bold=True)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # --- Table Header Row (starts from row 3 now) ---
    write_cell(ws, 3, 1, "Date", bold=True)
    write_cell(ws, 3, 2, "Validity (UTC)", bold=True)
    write_cell(ws, 3, 3, "Forecast & Elements verified", bold=True)
    write_cell(ws, 3, 7, "Realised & Elements verified", bold=True)
    write_cell(ws, 3, 10, "Accuracy", bold=True)

    # Merge header cells
    ws.merge_cells('A3:A4')
    ws.merge_cells('B3:B4')
    ws.merge_cells('C3:F3')
    ws.merge_cells('G3:I3')
    ws.merge_cells('J3:L3')

    # Row 4 headers
    headers_bottom = [
        "", "", "FL", "Wind Direction", "Speed", "Temp.",
        "Wind Direction", "Speed (KT)", "Temp.",
        "Wind Dir", "Speed", "Temp"
    ]
    for col_num, val in enumerate(headers_bottom, start=1):
        write_cell(ws, 4, col_num, val, bold=True)

    # --- Data Rows start at row 5 ---
    current_row = 5

    altitude_to_fl = {
    3000: "FL 100 (3000 M)",
    2100: "FL 070 (2100 M)",
    1500: "FL 050 (1500 M)",
    900:  "FL 030 (900 M)",
    600:  "FL 020 (600 M)",
    300:  "FL 010 (300 M)"
    }

    fl_labels = list(altitude_to_fl.values())
    filtered_rows = [row for row in data_rows if row.get("fl") in fl_labels]

    overall_temp_acc = round(sum(1 for r in data_rows if r.get("temp_acc") == "CORRECT") / len(data_rows) * 100, 2) if data_rows else 0
    overall_wind_acc = round(sum(1 for r in data_rows if r.get("speed_acc") == "CORRECT") / len(data_rows) * 100, 2) if data_rows else 0
    overall_wind_dir_acc = round(sum(1 for r in data_rows if r.get("wind_dir_acc") == "CORRECT") / len(data_rows) * 100, 2) if data_rows else 0

# For FL accuracy (only FL levels)
    fl_temp_acc = round(sum(1 for r in filtered_rows if r.get("temp_acc") == "CORRECT") / len(filtered_rows) * 100, 2) if filtered_rows else None
    fl_wind_acc = round(sum(1 for r in filtered_rows if r.get("speed_acc") == "CORRECT") / len(filtered_rows) * 100, 2) if filtered_rows else None
    fl_wind_dir_acc = round(sum(1 for r in filtered_rows if r.get("wind_dir_acc") == "CORRECT") / len(filtered_rows) * 100, 2) if filtered_rows else None


    for idx, row in enumerate(filtered_rows):
        row_key = f"{row.get('date', '')}_{row.get('validity', '')}"
        print(f"Checking weather_info for key: {row_key}")
        if weather_info:
            print("Available keys in weather_info:", weather_info.keys())

        values = [
            row.get("date", ""),
            row.get("validity", ""),
            row.get("fl", ""),
            row.get("forecast_wind_dir", ""),
            row.get("forecast_speed", ""),
            row.get("forecast_temp", ""),
            row.get("actual_wind_dir", ""),
            row.get("actual_speed", ""),
            f"{float(row.get('actual_temp', 0)):.2f}",
            row.get("wind_dir_acc", ""),
            row.get("speed_acc", ""),
            row.get("temp_acc", ""),
            # row.get("fl_accuracy_summary", {}),
        ]
        for col_num, val in enumerate(values, start=1):
            write_cell(ws, current_row, col_num, val)
        current_row += 1
        print(values)

        # Check next row key
        next_row_key = None
        if idx + 1 < len(data_rows):
            next_row = data_rows[idx + 1]
            next_row_key = f"{next_row.get('date', '')}_{next_row.get('validity', '')}"

        if row_key != next_row_key or idx == len(filtered_rows) - 1:
            weather_forecast = weather_info.get(row_key, {}).get("weather_forecast", "") if weather_info else ""
            weather_realised = " / ".join(weather_info.get(row_key, {}).get("matched", [])) if weather_info else ""
            weather_accuracy = weather_info.get(row_key, {}).get("accuracy", "") if weather_info else ""

            print(f"Writing weather row for {row_key}")  # DEBUG
            weather_row = [
                row.get("date", ""),
                row.get("validity", ""),
                "Significant Weather",
                "", "", weather_forecast,
                "", "", weather_realised,
                "", "", weather_accuracy
            ]
            for col_num, val in enumerate(weather_row, start=1):
                write_cell(ws, current_row, col_num, val)
            current_row += 1

    
    current_row += 1
  # or filter differently if needed

# Calculate overall accuracies
    ws.cell(row=current_row, column=1, value=f"Overall Temperature Accuracy: {overall_temp_acc}%")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1

    ws.cell(row=current_row, column=1, value=f"Overall Wind Speed Accuracy: {overall_wind_acc}%")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1

    ws.cell(row=current_row, column=1, value=f"Overall Wind Direction Accuracy: {overall_wind_dir_acc}%")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="Accuracy Summary for FL 010 (300 M) to FL 100 (3000 M)")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1

# Table header
    headers = ["FL Range", "Temperature Accuracy (%)", "Wind Speed Accuracy (%)", "Wind Direction Accuracy (%)"]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=current_row, column=col_num, value=header)
        ws.cell(row=current_row, column=col_num).font = Font(bold=True)
    current_row += 1

# Single summary row for all FLs
    ws.cell(row=current_row, column=1, value="FL 010 (300 M) to FL 100 (3000 M)")
    ws.cell(row=current_row, column=2, value=fl_temp_acc)
    ws.cell(row=current_row, column=3, value=fl_wind_acc)
    ws.cell(row=current_row, column=4, value=fl_wind_dir_acc)
    # ws.cell(row=current_row, column=5, value=fl_accuracy_summary["Total Levels"])
    current_row += 1

    # Leave a blank row after FL summary
    

    # Auto adjust widths
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    wb.save(file_path)
    return file_path

